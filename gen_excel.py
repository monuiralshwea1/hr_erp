"""
gen_excel.py - Build Excel from local JSON export
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JSON_PATH = r"C:\Users\ebh\Desktop\erp\stage\hr_erp_app\_export.json"
OUT_PATH = r"C:\Users\ebh\Desktop\erp\stage\hr_erp_app\shahadhi_full_report.xlsx"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def write_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border


def write_row(ws, row, values):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")


def main():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()

    # ── Sheet 1: الموظفين ──
    ws1 = wb.active
    ws1.title = "الموظفين"
    h1 = ["رقم الموظف", "اسم الموظف", "الجوال", "القسم", "الفرع", "تاريخ الالتحاق", "الحالة"]
    write_header(ws1, 1, h1)
    for i, emp in enumerate(data["employees"], 2):
        write_row(ws1, i, [
            emp["name"], emp["employee_name"], emp.get("cell_number", ""),
            emp.get("department", ""), emp.get("branch", ""),
            str(emp.get("date_of_joining", "")), emp.get("status", ""),
        ])
    for col in range(1, len(h1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 20

    # ── Sheet 2: الأقسام والفرع ──
    ws2 = wb.create_sheet("الأقسام والفرع")
    write_header(ws2, 1, ["القسم", "عدد الموظفين"])
    dept_counts = {}
    for emp in data["employees"]:
        d = emp.get("department", "غير محدد")
        dept_counts[d] = dept_counts.get(d, 0) + 1
    for i, (dept, cnt) in enumerate(sorted(dept_counts.items(), key=lambda x: -x[1]), 2):
        write_row(ws2, i, [dept, cnt])
    br = len(dept_counts) + 4
    write_header(ws2, br, ["الفرع", "عدد الموظفين"])
    branch_counts = {}
    for emp in data["employees"]:
        b = emp.get("branch", "غير محدد")
        branch_counts[b] = branch_counts.get(b, 0) + 1
    for i, (bname, cnt) in enumerate(sorted(branch_counts.items(), key=lambda x: -x[1]), br + 1):
        write_row(ws2, i, [bname, cnt])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    # ── Sheet 3: الرواتب والبدلات ──
    ws3 = wb.create_sheet("الرواتب والبدلات")
    h3 = ["رقم الموظف", "اسم الموظف", "الجوال", "القسم", "الفرع",
          "الراتب الاساسي", "بدل مواصلات 20%", "بدل مخاطر 50%",
          "بدل طبيعه عمل 60%", "بدل سكن 15%", "الاجمالي"]
    write_header(ws3, 1, h3)
    total_base = 0
    total_all = 0
    for i, ssa in enumerate(data["ssas"], 2):
        base = ssa.get("base", 0) or 0
        trn = round(base * 0.20)
        risk = round(base * 0.50)
        nat = round(base * 0.60)
        hou = round(base * 0.15)
        total = base + trn + risk + nat + hou
        total_base += base
        total_all += total
        write_row(ws3, i, [
            ssa["employee"], ssa["employee_name"], ssa.get("cell_number", ""),
            ssa.get("department", ""), ssa.get("branch", ""),
            base, trn, risk, nat, hou, total,
        ])
    total_row = len(data["ssas"]) + 3
    ws3.cell(row=total_row, column=5, value="الاجمالي").font = Font(bold=True)
    ws3.cell(row=total_row, column=6, value=total_base).font = Font(bold=True)
    ws3.cell(row=total_row, column=11, value=total_all).font = Font(bold=True)
    for col in range(1, len(h3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 18
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 30

    # ── Sheet 4: فترات الدوام ──
    ws4 = wb.create_sheet("فترات الدوام")
    h4 = ["نوع الدوام", "بداية الدوام", "نهاية الدوام", "الفترة", "من", "الى", "استراحة؟"]
    write_header(ws4, 1, h4)
    row = 2
    for s in data["shifts"]:
        write_row(ws4, row, [
            s["shift_name"], str(s.get("start_time", "")), str(s.get("end_time", "")),
            s.get("period_name", ""), str(s.get("p_start", "")),
            str(s.get("p_end", "")), "نعم" if s.get("is_break") else "لا",
        ])
        row += 1
    for col in range(1, len(h4) + 1):
        ws4.column_dimensions[get_column_letter(col)].width = 25

    # ── Sheet 5: ملخص الحضور ──
    ws5 = wb.create_sheet("ملخص الحضور")
    emp_att = {}
    for a in data["att_summary"]:
        emp = a["employee"]
        if emp not in emp_att:
            emp_att[emp] = {
                "name": emp, "name2": a["employee_name"],
                "branch": a.get("branch", ""), "dept": a.get("department", ""),
                "shift": a.get("shift", ""),
                "Present": 0, "Absent": 0, "Half Day": 0, "On Leave": 0,
                "total_hours": 0, "late_days": 0,
            }
        st = a.get("status", "")
        if st in emp_att[emp]:
            emp_att[emp][st] += a.get("cnt", 0)
        emp_att[emp]["total_hours"] += a.get("total_hours", 0) or 0
        emp_att[emp]["late_days"] += a.get("late_days", 0) or 0

    h5 = ["رقم الموظف", "اسم الموظف", "الفرع", "القسم", "نوع الدوام",
          "ايام الحضور", "ايام الغياب", "نصف يوم", "اجمالي الساعات", "ايام التأخير"]
    write_header(ws5, 1, h5)
    for i, (eid, att) in enumerate(sorted(emp_att.items()), 2):
        write_row(ws5, i, [
            att["name"], att["name2"], att["branch"], att["dept"], att["shift"],
            att["Present"], att["Absent"], att["Half Day"],
            round(att["total_hours"], 1), att["late_days"],
        ])
    for col in range(1, len(h5) + 1):
        ws5.column_dimensions[get_column_letter(col)].width = 18
    ws5.column_dimensions["A"].width = 20
    ws5.column_dimensions["B"].width = 30

    # ── Sheet 6: تفاصيل الفترات ──
    ws6 = wb.create_sheet("تفاصيل الفترات")
    h6 = ["رقم الموظف", "اسم الموظف", "نوع الدوام", "الفترة",
          "الحالة", "العدد", "الساعات", "دقائق التأخير", "ساعات الغياب"]
    write_header(ws6, 1, h6)
    row = 2
    for pd in data["pd_emp"]:
        write_row(ws6, row, [
            pd["employee"], pd.get("employee_name", ""),
            pd.get("shift_type", ""), pd.get("period_name", ""),
            pd.get("period_status", ""), pd.get("cnt", 0),
            round(pd.get("hours", 0) or 0, 1),
            round(pd.get("late_min", 0) or 0, 1),
            round(pd.get("absent_hrs", 0) or 0, 1),
        ])
        row += 1
    for col in range(1, len(h6) + 1):
        ws6.column_dimensions[get_column_letter(col)].width = 18
    ws6.column_dimensions["B"].width = 30

    # ── Sheet 7: ملخص الفترات ──
    ws7 = wb.create_sheet("ملخص الفترات")
    h7 = ["الحالة", "العدد", "اجمالي الساعات", "اجمالي التأخير (دقائق)"]
    write_header(ws7, 1, h7)
    for i, ps in enumerate(data["pd_summary"], 2):
        write_row(ws7, i, [
            ps.get("period_status", ""),
            ps.get("cnt", 0),
            round(ps.get("total_hours", 0) or 0, 1),
            round(ps.get("total_late", 0) or 0, 1),
        ])
    for col in range(1, len(h7) + 1):
        ws7.column_dimensions[get_column_letter(col)].width = 25

    # ── Sheet 8: كشوف الرواتب ──
    ws8 = wb.create_sheet("كشوف الرواتب")
    h8 = ["رقم الموظف", "اسم الموظف", "القسم", "الفرع", "من", "الى",
          "الاجمالي", "الخصومات", "الصافي", "الحالة"]
    write_header(ws8, 1, h8)
    row = 2
    for ss in data.get("salary_slips", []):
        write_row(ws8, row, [
            ss["employee"], ss.get("employee_name", ""),
            ss.get("department", ""), ss.get("branch", ""),
            str(ss.get("start_date", "")), str(ss.get("end_date", "")),
            ss.get("gross_pay", 0),
            ss.get("total_deduction", 0), ss.get("net_pay", 0), ss.get("status", ""),
        ])
        row += 1
    for col in range(1, len(h8) + 1):
        ws8.column_dimensions[get_column_letter(col)].width = 18
    ws8.column_dimensions["B"].width = 30

    # ── Sheet 9: قيود الرواتب ──
    ws9 = wb.create_sheet("قيود الرواتب")
    h9 = ["رقم القيد", "من", "الى", "عدد الموظفين", "اجمالي الرواتب"]
    write_header(ws9, 1, h9)
    for i, pe in enumerate(data.get("payroll_entries", []), 2):
        write_row(ws9, i, [
            pe["name"], str(pe.get("start_date", "")),
            str(pe.get("end_date", "")), pe.get("number_of_employees", 0),
            0,
        ])
    for col in range(1, len(h9) + 1):
        ws9.column_dimensions[get_column_letter(col)].width = 22

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Employees: {len(data['employees'])}")
    print(f"SSAs: {len(data['ssas'])}")
    print(f"Salary Slips: {len(data.get('salary_slips', []))}")
    print(f"Payroll Entries: {len(data.get('payroll_entries', []))}")
    print(f"Total base salary: {total_base:,}")
    print(f"Total with allowances: {total_all:,}")


if __name__ == "__main__":
    main()
